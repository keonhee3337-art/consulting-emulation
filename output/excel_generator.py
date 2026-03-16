"""
excel_generator.py — Export DCF + Comps valuation model to Excel.

Sheets:
  1. Summary      — Valuation range, key metrics
  2. DCF Model    — Year-by-year FCF projection, PV, terminal value
  3. Comps        — Peer EV/EBITDA table, target application
  4. Financials   — Historical revenue / profit / margins

Usage:
    from output.excel_generator import generate_excel
    buf = generate_excel(company_name, dcf_data, comps_data, financial_rows)
    # buf is an io.BytesIO — pass directly to st.download_button
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Palette ────────────────────────────────────────────────────────────────
NAVY_HEX   = "1B2A4A"
GOLD_HEX   = "C9A84C"
LIGHT_HEX  = "EAF0F6"
WHITE_HEX  = "FFFFFF"
GREY_HEX   = "F5F5F5"
RED_HEX    = "DC2626"
GREEN_HEX  = "059669"

# ── DCF constants (mirror valuation_agent.py) ──────────────────────────────
WACC             = 0.10
TERMINAL_GROWTH  = 0.025
PROJECTION_YEARS = 5
FCF_MARGIN       = 0.85

# ── Style helpers ──────────────────────────────────────────────────────────
def _hdr(bold=True, color=WHITE_HEX, bg=NAVY_HEX, size=11):
    return {
        "font":      Font(bold=bold, color=color, size=size, name="Calibri"),
        "fill":      PatternFill("solid", fgColor=bg),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def _label(bold=True, size=10):
    return {
        "font":      Font(bold=bold, size=size, name="Calibri"),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }

def _num(size=10):
    return {
        "font":      Font(size=size, name="Calibri"),
        "alignment": Alignment(horizontal="right", vertical="center"),
    }

def _apply(cell, styles: dict):
    for attr, val in styles.items():
        setattr(cell, attr, val)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _section_title(ws, row, col, text, colspan=6):
    cell = ws.cell(row=row, column=col, value=text)
    _apply(cell, _hdr())
    if colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row,   end_column=col + colspan - 1
        )
    return row + 1

def _kv(ws, row, key, value, fmt=None, indent=0):
    lbl = ws.cell(row=row, column=1, value=(" " * indent) + key)
    _apply(lbl, _label())
    val = ws.cell(row=row, column=2, value=value)
    _apply(val, _num())
    if fmt:
        val.number_format = fmt
    return row + 1

def _spacer(ws, row):
    return row + 1


# ── Sheet 1: Summary ───────────────────────────────────────────────────────
def _sheet_summary(wb, company, dcf, comps):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title
    title = ws.cell(row=1, column=1, value=f"Valuation Summary — {company}")
    _apply(title, _hdr(size=14))
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    sub = ws.cell(row=2, column=1,
                  value=f"Generated: {date.today().strftime('%Y-%m-%d')}  |  DART financial data  |  KRW billions")
    _apply(sub, {"font": Font(size=9, color="888888", name="Calibri"),
                 "alignment": Alignment(horizontal="left")})
    ws.merge_cells("A2:F2")

    r = 4
    # ── Valuation Range ──────────────────────────────────────────────────
    r = _section_title(ws, r, 1, "VALUATION RANGE", colspan=6)

    ev_dcf   = dcf.get("ev_dcf_bn")
    ev_comps = comps.get("ev_comps_bn")
    midpoint = round((ev_dcf + ev_comps) / 2, 1) if ev_dcf and ev_comps else None
    low      = min(ev_dcf, ev_comps) if ev_dcf and ev_comps else None
    high     = max(ev_dcf, ev_comps) if ev_dcf and ev_comps else None

    r = _kv(ws, r, "DCF Enterprise Value (KRW bn)",   ev_dcf,    '#,##0.0')
    r = _kv(ws, r, "EV/EBITDA Comps (KRW bn)",        ev_comps,  '#,##0.0')
    r = _kv(ws, r, "Midpoint (KRW bn)",                midpoint,  '#,##0.0')
    r = _kv(ws, r, "Low — High (KRW bn)",
            f"{low:,.1f} — {high:,.1f}" if low else "N/A")

    r = _spacer(ws, r)
    # ── DCF Snapshot ─────────────────────────────────────────────────────
    r = _section_title(ws, r, 1, "DCF SNAPSHOT", colspan=6)
    r = _kv(ws, r, "Base FCF (KRW bn)",              dcf.get("fcf_base"),     '#,##0.0')
    r = _kv(ws, r, "FCF Growth Rate Applied",        dcf.get("growth_rate"),  '0.0%')
    r = _kv(ws, r, "PV of Projected FCFs (KRW bn)",  dcf.get("pv_sum_fcfs"),  '#,##0.0')
    r = _kv(ws, r, "PV of Terminal Value (KRW bn)",  dcf.get("pv_terminal"),  '#,##0.0')
    r = _kv(ws, r, "WACC",                           WACC,                    '0.0%')
    r = _kv(ws, r, "Terminal Growth Rate",           TERMINAL_GROWTH,         '0.0%')

    r = _spacer(ws, r)
    # ── Comps Snapshot ────────────────────────────────────────────────────
    r = _section_title(ws, r, 1, "COMPS SNAPSHOT", colspan=6)
    r = _kv(ws, r, "Target EBITDA Proxy (KRW bn)",  comps.get("target_ebitda"),   '#,##0.0')
    r = _kv(ws, r, "Peer Median EV/EBITDA Multiple", comps.get("median_multiple"), '0.0"x"')
    r = _kv(ws, r, "EV via Comps (KRW bn)",         ev_comps,                     '#,##0.0')

    r = _spacer(ws, r)
    r = _section_title(ws, r, 1, "ANALYST NOTE", colspan=6)
    note = ws.cell(row=r, column=1,
        value="EBITDA proxy = Operating Profit (D&A excluded). "
              "Peer universe = 3 Korean electronics companies (cross-sector — interpret cautiously). "
              "EV figures exclude net debt — do not represent equity value.")
    note.font = Font(size=9, italic=True, color="555555", name="Calibri")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{r}:F{r}")
    ws.row_dimensions[r].height = 40

    _set_col_width(ws, 1, 42)
    _set_col_width(ws, 2, 18)
    for c in range(3, 7):
        _set_col_width(ws, c, 12)


# ── Sheet 2: DCF Model ─────────────────────────────────────────────────────
def _sheet_dcf(wb, company, dcf, financial_rows):
    ws = wb.create_sheet("DCF Model")
    ws.sheet_view.showGridLines = False

    title = ws.cell(row=1, column=1, value=f"DCF Model — {company}")
    _apply(title, _hdr(size=13))
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 26

    r = 3
    # ── Assumptions block ─────────────────────────────────────────────────
    r = _section_title(ws, r, 1, "ASSUMPTIONS", colspan=8)
    for key, val, fmt in [
        ("WACC",                  WACC,             "0.0%"),
        ("Terminal Growth Rate",  TERMINAL_GROWTH,  "0.0%"),
        ("Projection Period",     PROJECTION_YEARS, '0" years"'),
        ("FCF Proxy",             "Net Income × 0.85", None),
        ("Base FCF (KRW bn)",     dcf.get("fcf_base"),     "#,##0.0"),
        ("FCF Growth Rate",       dcf.get("growth_rate"),  "0.0%"),
        ("Growth Note",           dcf.get("growth_note"),  None),
    ]:
        lbl = ws.cell(row=r, column=1, value=key)
        _apply(lbl, _label())
        v = ws.cell(row=r, column=2, value=val)
        _apply(v, _num())
        if fmt:
            v.number_format = fmt
        r += 1

    r = _spacer(ws, r)
    # ── Year-by-year FCF projection ────────────────────────────────────────
    r = _section_title(ws, r, 1, "YEAR-BY-YEAR FCF PROJECTION", colspan=8)

    # Header row
    headers = ["Year", "FCF (KRW bn)", "Discount Factor", "PV of FCF (KRW bn)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        _apply(cell, _hdr(bg=GOLD_HEX, color=NAVY_HEX))
    r += 1

    fcf_base    = dcf.get("fcf_base", 0)
    growth_rate = dcf.get("growth_rate", 0)
    last_year   = financial_rows[-1]["year"] if financial_rows else date.today().year

    for t in range(1, PROJECTION_YEARS + 1):
        fcf_t  = fcf_base * (1 + growth_rate) ** t
        disc_t = 1 / (1 + WACC) ** t
        pv_t   = fcf_t * disc_t

        vals = [last_year + t, fcf_t, disc_t, pv_t]
        fmts = ['0', '#,##0.0', '0.0000', '#,##0.0']
        bg = LIGHT_HEX if t % 2 == 0 else WHITE_HEX
        for ci, (v, f) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=ci, value=v)
            _apply(cell, _num())
            cell.number_format = f
            cell.fill = PatternFill("solid", fgColor=bg)
        r += 1

    r = _spacer(ws, r)
    # ── Terminal Value ────────────────────────────────────────────────────
    r = _section_title(ws, r, 1, "TERMINAL VALUE", colspan=8)

    fcf_y5 = fcf_base * (1 + growth_rate) ** PROJECTION_YEARS
    tv     = fcf_y5 * (1 + TERMINAL_GROWTH) / (WACC - TERMINAL_GROWTH)
    pv_tv  = tv / (1 + WACC) ** PROJECTION_YEARS

    for key, val, fmt in [
        ("FCF Year 5 (KRW bn)",            fcf_y5,                "#,##0.0"),
        ("Terminal Value — Gordon Growth (KRW bn)", tv,           "#,##0.0"),
        ("Discount Factor (Year 5)",       1 / (1 + WACC)**PROJECTION_YEARS, "0.0000"),
        ("PV of Terminal Value (KRW bn)",  pv_tv,                 "#,##0.0"),
        ("PV of Projected FCFs (KRW bn)",  dcf.get("pv_sum_fcfs"), "#,##0.0"),
        ("Enterprise Value — DCF (KRW bn)", dcf.get("ev_dcf_bn"),  "#,##0.0"),
    ]:
        lbl = ws.cell(row=r, column=1, value=key)
        _apply(lbl, _label(bold=(key.startswith("Enterprise"))))
        if key.startswith("Enterprise"):
            lbl.font = Font(bold=True, color=NAVY_HEX, size=11, name="Calibri")
        v = ws.cell(row=r, column=2, value=val)
        _apply(v, _num())
        if fmt:
            v.number_format = fmt
        if key.startswith("Enterprise"):
            v.font = Font(bold=True, color=NAVY_HEX, size=11, name="Calibri")
        r += 1

    _set_col_width(ws, 1, 40)
    for c in range(2, 9):
        _set_col_width(ws, c, 20)


# ── Sheet 3: Comps ─────────────────────────────────────────────────────────
def _sheet_comps(wb, company, comps):
    ws = wb.create_sheet("Comps (EV-EBITDA)")
    ws.sheet_view.showGridLines = False

    title = ws.cell(row=1, column=1, value=f"Comparable Company Analysis — {company}")
    _apply(title, _hdr(size=13))
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 26

    r = 3
    r = _section_title(ws, r, 1, "PEER UNIVERSE", colspan=5)

    # Header
    for ci, h in enumerate(["Company", "EBITDA Proxy (KRW bn)", "Implied EV (KRW bn)", "EV/EBITDA Multiple"], 1):
        cell = ws.cell(row=r, column=ci, value=h)
        _apply(cell, _hdr(bg=GOLD_HEX, color=NAVY_HEX))
    r += 1

    for i, peer in enumerate(comps.get("peer_details", [])):
        ebitda   = peer["ebitda"]
        multiple = peer["multiple"]
        impl_ev  = ebitda * multiple
        bg = LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        for ci, (v, f) in enumerate([
            (peer["peer"], None),
            (ebitda,       "#,##0.0"),
            (impl_ev,      "#,##0.0"),
            (multiple,     '0.0"x"'),
        ], 1):
            cell = ws.cell(row=r, column=ci, value=v)
            _apply(cell, _num() if ci > 1 else _label(bold=False))
            cell.fill = PatternFill("solid", fgColor=bg)
            if f:
                cell.number_format = f
        r += 1

    # Median row
    med_cell = ws.cell(row=r, column=1, value="Peer Median")
    _apply(med_cell, _label(bold=True))
    med_cell.fill = PatternFill("solid", fgColor=NAVY_HEX)
    med_cell.font = Font(bold=True, color=WHITE_HEX, name="Calibri")
    mv = ws.cell(row=r, column=4, value=comps.get("median_multiple"))
    _apply(mv, _num())
    mv.number_format = '0.0"x"'
    mv.font = Font(bold=True, color=WHITE_HEX, name="Calibri")
    mv.fill = PatternFill("solid", fgColor=NAVY_HEX)
    r += 2

    r = _section_title(ws, r, 1, "TARGET APPLICATION", colspan=5)
    r = _kv(ws, r, f"Target: {company} EBITDA proxy (KRW bn)",
            comps.get("target_ebitda"), "#,##0.0")
    r = _kv(ws, r, "× Peer Median EV/EBITDA",
            comps.get("median_multiple"), '0.0"x"')
    r = _kv(ws, r, "= Enterprise Value via Comps (KRW bn)",
            comps.get("ev_comps_bn"), "#,##0.0")

    # Bold the EV result
    ws.cell(row=r-1, column=1).font = Font(bold=True, color=NAVY_HEX, name="Calibri")
    ws.cell(row=r-1, column=2).font = Font(bold=True, color=NAVY_HEX, name="Calibri")

    r = _spacer(ws, r)
    note = ws.cell(row=r, column=1,
        value="Note: EBITDA proxy = Operating Profit. "
              "Peer EV implied via perpetuity midpoint (0% / 2.5% growth). "
              "Cross-sector peer universe — interpret multiples cautiously.")
    note.font = Font(size=9, italic=True, color="555555", name="Calibri")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(f"A{r}:E{r}")
    ws.row_dimensions[r].height = 36

    _set_col_width(ws, 1, 36)
    for c in range(2, 6):
        _set_col_width(ws, c, 22)


# ── Sheet 4: Financials ────────────────────────────────────────────────────
def _sheet_financials(wb, company, rows):
    ws = wb.create_sheet("Financials")
    ws.sheet_view.showGridLines = False

    title = ws.cell(row=1, column=1, value=f"Historical Financials — {company}  (KRW billions)")
    _apply(title, _hdr(size=13))
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 26

    r = 3
    headers = ["Year", "Revenue", "Operating Profit", "Net Profit",
               "Operating Margin %", "Net Margin %", "FCF Proxy (Net×0.85)"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        _apply(cell, _hdr(bg=GOLD_HEX, color=NAVY_HEX))
    r += 1

    for i, row in enumerate(rows):
        rev  = row.get("revenue_billion_krw")
        op   = row.get("operating_profit_billion_krw")
        net  = row.get("net_profit_billion_krw")
        op_m = round(op / rev * 100, 1) if rev and op is not None else None
        net_m = round(net / rev * 100, 1) if rev and net is not None else None
        fcf_p = round(net * FCF_MARGIN, 1) if net is not None else None

        bg = LIGHT_HEX if i % 2 == 0 else WHITE_HEX
        vals_fmts = [
            (row["year"],      "0"),
            (rev,              "#,##0.0"),
            (op,               "#,##0.0"),
            (net,              "#,##0.0"),
            (op_m,             '0.0"%"'),
            (net_m,            '0.0"%"'),
            (fcf_p,            "#,##0.0"),
        ]
        for ci, (v, f) in enumerate(vals_fmts, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            _apply(cell, _num())
            if f:
                cell.number_format = f
            cell.fill = PatternFill("solid", fgColor=bg)
            # Colour negative net profit red
            if ci in (4, 6) and isinstance(v, (int, float)) and v is not None and v < 0:
                cell.font = Font(color=RED_HEX, size=10, name="Calibri")
        r += 1

    _set_col_width(ws, 1, 8)
    for c in range(2, 8):
        _set_col_width(ws, c, 22)


# ── Public API ─────────────────────────────────────────────────────────────
def generate_excel(
    company_name: str,
    dcf_data: dict,
    comps_data: dict,
    financial_rows: list[dict],
) -> io.BytesIO:
    """
    Build a multi-sheet Excel workbook and return it as a BytesIO buffer.
    Pass the buffer directly to st.download_button(data=buf).
    """
    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    _sheet_summary(wb, company_name, dcf_data, comps_data)
    _sheet_dcf(wb, company_name, dcf_data, financial_rows)
    _sheet_comps(wb, company_name, comps_data)
    _sheet_financials(wb, company_name, financial_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
